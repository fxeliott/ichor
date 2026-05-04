"""AI-GPR Index daily collector — LLM-scored geopolitical risk.

Caldara & Iacoviello (Federal Reserve Board) maintain two indices :

  - **GPR** (classic) : monthly, since 1985, based on newspaper text searches
  - **AI-GPR** (NEW since 2024) : **daily**, since 1960, LLM-scored from
    expanded text corpus + LLM classification

The AI-GPR index is more reactive — it can spike in days vs months for
GPR. Useful for our session cards because geopolitical premium can drive
intraday gold + USD haven moves.

Data is published as CSV / Excel at :
  https://www.matteoiacoviello.com/ai_gpr.html

This collector downloads the latest CSV, returns the time series. Cached
heavily because the file is updated at most once per day.

License : academic, free for non-commercial use. Commercial inquiries
to the authors.
"""

from __future__ import annotations

import csv
import io
import sys
from dataclasses import dataclass
from datetime import date, datetime, timezone

import httpx
import structlog

log = structlog.get_logger(__name__)

AI_GPR_CSV_URL = "https://www.matteoiacoviello.com/gpr_files/data_gpr_daily_recent.xls"
AI_GPR_HOMEPAGE = "https://www.matteoiacoviello.com/ai_gpr.html"


@dataclass(frozen=True)
class AiGprObservation:
    """One daily AI-GPR reading."""

    observation_date: date
    """Calendar day of the reading."""

    ai_gpr: float
    """AI-GPR Index value. Higher = more geopolitical risk."""

    fetched_at: datetime


def _is_xls_binary(body: bytes) -> bool:
    """Detect Microsoft CFB binary signature (D0CF11E0A1B11AE1) — the
    legacy .xls format that csv.DictReader cannot parse."""
    return body[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _is_xlsx_zip(body: bytes) -> bool:
    """Detect ZIP signature (PK\x03\x04) — the .xlsx Office 2007+ format."""
    return body[:4] == b"PK\x03\x04"


def _parse_xls(body: bytes) -> list[AiGprObservation]:
    """Parse legacy .xls (CFB) via xlrd. Returns [] on any error."""
    try:
        import xlrd  # type: ignore[import-untyped]
    except ImportError:
        log.warning("ai_gpr.xlrd_unavailable")
        return []
    try:
        wb = xlrd.open_workbook(file_contents=body)
        sheet = wb.sheet_by_index(0)
        header = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
        date_idx = next(
            (i for i, h in enumerate(header) if h.lower() in {"date", "day"}),
            None,
        )
        value_idx = next(
            (i for i, h in enumerate(header)
             if h.upper() in {"AI_GPR", "GPR_DAILY", "GPR", "GPRD"}),
            None,
        )
        if date_idx is None or value_idx is None:
            log.warning("ai_gpr.xls_header_unrecognized", header=header)
            return []
        fetched = datetime.now(timezone.utc)
        out: list[AiGprObservation] = []
        for row_idx in range(1, sheet.nrows):
            try:
                d_raw = sheet.cell_value(row_idx, date_idx)
                v_raw = sheet.cell_value(row_idx, value_idx)
                if v_raw == "" or v_raw is None:
                    continue
                if isinstance(d_raw, float):
                    d_tuple = xlrd.xldate_as_tuple(d_raw, wb.datemode)
                    bar_date = date(d_tuple[0], d_tuple[1], d_tuple[2])
                elif isinstance(d_raw, str):
                    s = d_raw.strip()
                    try:
                        bar_date = date.fromisoformat(s)
                    except ValueError:
                        bar_date = datetime.strptime(s, "%m/%d/%Y").date()
                else:
                    continue
                out.append(
                    AiGprObservation(
                        observation_date=bar_date,
                        ai_gpr=float(v_raw),
                        fetched_at=fetched,
                    )
                )
            except (TypeError, ValueError):
                continue
        return out
    except Exception as e:
        log.warning("ai_gpr.xls_parse_failed", error=str(e))
        return []


def _parse_xlsx(body: bytes) -> list[AiGprObservation]:
    """Parse .xlsx via openpyxl. Returns [] on any error."""
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError:
        log.warning("ai_gpr.openpyxl_unavailable")
        return []
    try:
        wb = load_workbook(filename=io.BytesIO(body), data_only=True, read_only=True)
        sheet = wb.active
        rows = sheet.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
        date_idx = next(
            (i for i, h in enumerate(header) if h.lower() in {"date", "day"}),
            None,
        )
        value_idx = next(
            (i for i, h in enumerate(header)
             if h.upper() in {"AI_GPR", "GPR_DAILY", "GPR", "GPRD"}),
            None,
        )
        if date_idx is None or value_idx is None:
            log.warning("ai_gpr.xlsx_header_unrecognized", header=header)
            return []
        fetched = datetime.now(timezone.utc)
        out: list[AiGprObservation] = []
        for row in rows:
            try:
                d_raw = row[date_idx]
                v_raw = row[value_idx]
                if v_raw is None or v_raw == "":
                    continue
                if isinstance(d_raw, datetime):
                    bar_date = d_raw.date()
                elif isinstance(d_raw, date):
                    bar_date = d_raw
                elif isinstance(d_raw, str):
                    s = d_raw.strip()
                    try:
                        bar_date = date.fromisoformat(s)
                    except ValueError:
                        bar_date = datetime.strptime(s, "%m/%d/%Y").date()
                else:
                    continue
                out.append(
                    AiGprObservation(
                        observation_date=bar_date,
                        ai_gpr=float(v_raw),
                        fetched_at=fetched,
                    )
                )
            except (TypeError, ValueError):
                continue
        return out
    except Exception as e:
        log.warning("ai_gpr.xlsx_parse_failed", error=str(e))
        return []


def _parse_csv(body: bytes) -> list[AiGprObservation]:
    """Detect upstream format and dispatch ; falls back to CSV.

    Order of magic-byte checks :
      1. CFB (.xls vintage)        → xlrd
      2. ZIP (.xlsx Office 2007+)  → openpyxl
      3. text/csv                  → DictReader (canonical path)
    """
    if _is_xls_binary(body):
        return _parse_xls(body)
    if _is_xlsx_zip(body):
        return _parse_xlsx(body)

    csv.field_size_limit(sys.maxsize)
    text = body.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    fields = set(reader.fieldnames or [])

    # Accept either historical naming (`GPR_DAILY`) or the newer `AI_GPR`.
    value_col = (
        "AI_GPR" if "AI_GPR" in fields
        else "GPR_DAILY" if "GPR_DAILY" in fields
        else None
    )
    date_col = (
        "date" if "date" in fields
        else "DATE" if "DATE" in fields
        else None
    )
    if value_col is None or date_col is None:
        log.warning("ai_gpr.unexpected_header", header=reader.fieldnames)
        return []

    fetched = datetime.now(timezone.utc)
    out: list[AiGprObservation] = []
    for row in reader:
        try:
            d = row[date_col].strip()
            v = row[value_col].strip()
            if not d or not v or v.lower() in {"na", "nan", "."}:
                continue
            # Try ISO first, then US date
            try:
                bar_date = date.fromisoformat(d)
            except ValueError:
                bar_date = datetime.strptime(d, "%m/%d/%Y").date()
            out.append(
                AiGprObservation(
                    observation_date=bar_date,
                    ai_gpr=float(v),
                    fetched_at=fetched,
                )
            )
        except (KeyError, ValueError) as e:
            log.warning("ai_gpr.parse_row_failed", error=str(e))
            continue
    return out


async def fetch_latest(*, timeout: float = 30.0) -> list[AiGprObservation]:
    """Download the latest AI-GPR CSV and return the parsed series.

    The site occasionally renames the file ; we try the canonical URL
    first and fall back to a homepage scrape if needed.
    """
    async with httpx.AsyncClient() as client:
        try:
            r = await client.get(
                AI_GPR_CSV_URL,
                timeout=timeout,
                follow_redirects=True,
                headers={"User-Agent": "IchorAIGPRCollector/0.1"},
            )
            r.raise_for_status()
        except httpx.HTTPError as e:
            log.warning("ai_gpr.fetch_failed", url=AI_GPR_CSV_URL, error=str(e))
            return []

    rows = _parse_csv(r.content)
    if not rows:
        log.warning("ai_gpr.empty_or_unparseable")
    rows.sort(key=lambda x: x.observation_date)
    return rows


def latest_n_days(rows: list[AiGprObservation], n: int = 30) -> list[AiGprObservation]:
    return rows[-n:] if len(rows) > n else rows


def delta_30d(rows: list[AiGprObservation]) -> float | None:
    """Z-score-style delta : current - mean(last 30) / std(last 30)."""
    import statistics
    if len(rows) < 30:
        return None
    window = [r.ai_gpr for r in rows[-30:]]
    current = window[-1]
    mean = sum(window) / len(window)
    try:
        std = statistics.stdev(window)
    except statistics.StatisticsError:
        return None
    return (current - mean) / std if std > 0 else 0.0
